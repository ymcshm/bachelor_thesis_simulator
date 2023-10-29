import os
import csv
import math
from statistics import mean,stdev
from tqdm import tqdm
import datetime
import openpyxl as px
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill,Font

def abc_from_number(number):#可視化するのに使う関数
    alphabet = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    colname = ""
    divend = number
    while divend > 0:
        modulo = (divend - 1) % 26
        colname = alphabet[modulo] + colname
        divend = (divend - modulo) // 26
    return colname

def colorBarRGB(car_id):#可視化するのに使う関数
    car_id=car_id%10
    if car_id==1:
        color="FF0000"#赤
    elif car_id==2:
        color="FFA500"#オレンジ
    elif car_id==3:
        color="00FF00"#黄緑
    elif car_id==4:
        color="007400"#緑
    elif car_id==5:
        color="00FFFF"#水色
    elif car_id==6:
        color="0000FF"#青
    elif car_id==7:
        color="8D0093"#紫
    elif car_id==8:
        color="FF00FF"#ピンク
    elif car_id==9:
        color="800000"#茶色
    elif car_id==0:
        color="808080"#グレー
    return (color)

def create_excel_file():
    wb = px.Workbook()
    ws = wb.active
    return wb,ws

def create_info_sheet(ws,time_max,car_max,road_list,prc_time):

    row_num = 1
    
    def write_info(row_num,title,value):
        ws.cell(row = row_num, column = 1).value=(title)
        ws.cell(row = row_num, column = 2).value=(value)
        row_num += 1
        return row_num

    row_num = write_info(row_num,"プログラム処理時間", prc_time)
    row_num = write_info(row_num,"シミュレーション時間[s]", time_max/10)
    row_num = write_info(row_num,"全交通量", car_max)

    for road_id,road in enumerate(road_list):
        for q_id,q_lane in enumerate(road.queue):
            row_num = write_info(row_num,"road_"+str(road_id)+"_"+str(q_id)+" 交通量",q_lane)

        for type_i,ratio in enumerate(road.ratio_type):
            row_num = write_info(row_num,"road_"+str(road_id)+" type"+str(type_i)+"割合",ratio)

        for type_i,sum in enumerate(road.sum_type):
            row_num = write_info(row_num,"road_"+str(road_id)+" type"+str(type_i)+"台数",sum)

        for ic_i,sum in enumerate(road.sum_ic_out_car):
            row_num = write_info(row_num,"road_"+str(road_id)+" IC"+str(ic_i)+"離脱台数",sum)

        row_num = write_info(row_num,"road_"+str(road_id)+" アグレッシブドライバー数",road.sum_ag_driver)
        row_num = write_info(row_num,"road_"+str(road_id)+" 迂回処理時間[s]",road.detour_time_count/10)

def detour_info_sheet(wb,car_log,traveltime_list,road_list,time_max,generate_time_max,interval):
    title = "迂回情報"
    ws = wb.create_sheet(title=title)

    #init
    highway = road_list[0]
    general = road_list[1]

    #迂回した車両idを記録
    ws.cell(row = 1, column = 1).value=("迂回車両id一覧")

    row_num = 2
    for detour in highway.detour_list:

        ws.cell(row = row_num, column = 1).value=(detour)
        row_num += 1

    #interval毎に範囲内車両の速度を記録する
    highway_start = 3500
    highway_end   = 4500
    highway_vel_list = []

    detour_start  = 1500
    detour_end    = 2500
    detour_vel_list  = []

    for car in car_log:
        
        if (car.time%interval == 0 and
            car.lane    != 0
            ):

            #高速道路の速度測定
            if (car.road_id == 0 and
                car.front > highway_start and
                car.front < highway_end
                ):

                highway_vel_list.append(car.vel*3.6)

            #一般道路の速度測定
            if (car.road_id == 1 and
                car.front > detour_start and
                car.front < detour_end
                ):

                detour_vel_list.append(car.vel*3.6)

    #高速道路の平均速度計算
    highway_ave_vel = mean(highway_vel_list)
    highway_ave_std = stdev(highway_vel_list)

    #迂回道路の平均速度計算
    detour_ave_vel  = mean(detour_vel_list)
    detour_ave_std  = stdev(detour_vel_list)

    #シート出力
    ws.cell(row = 1, column = 3).value=("3500-4500高速道路")
    ws.cell(row = 2, column = 3).value=("発生時間交通量")
    ws.cell(row = 3, column = 3).value=("平均速度")
    ws.cell(row = 4, column = 3).value=("標準偏差")

    #NOTE 時間交通量はとりあえず走行車線のみ表示(走行車線も追い越し車線ともに同じなので)
    ws.cell(row = 2, column = 4).value=(highway.queue[1]*3600/generate_time_max*10)
    ws.cell(row = 3, column = 4).value=(highway_ave_vel)
    ws.cell(row = 4, column = 4).value=(highway_ave_std)

    ws.cell(row = 5, column = 3).value=("1500-2500一般道路")
    ws.cell(row = 6, column = 3).value=("発生時間交通量")
    ws.cell(row = 7, column = 3).value=("平均速度")
    ws.cell(row = 8, column = 3).value=("標準偏差")

    #NOTE 時間交通量はとりあえず走行車線のみ表示(走行車線も追い越し車線ともに同じなので)
    ws.cell(row = 6, column = 4).value=(general.queue[1]*3600/generate_time_max*10)
    ws.cell(row = 7, column = 4).value=(detour_ave_vel)
    ws.cell(row = 8, column = 4).value=(detour_ave_std)

    #迂回台数カウント
    ws.cell(row = 10, column = 3).value=("一般道路発生時間交通量")
    ws.cell(row = 11, column = 3).value=("迂回台数")

    ws.cell(row = 10, column = 4).value=(general.queue[1]*3600/generate_time_max*10)
    ws.cell(row = 11, column = 4).value=(len(highway.detour_list))

    #制御開始時間表示
    ws.cell(row = 13, column = 3).value=("迂回制御開始時間")
    
    ws.cell(row = 13, column = 4).value=(highway.start_control_time)

    #平均旅行時間計算
    def average_traveltime(traveltime_list,detour_info,start_time=0):
        """
        平均旅行時間を計算する
        start_timeに平均を測定し始める消滅時間を設定する
        300を指定した場合、消滅時間300秒以降の車両の平均旅行時間を計算する。
        """
        tmp_traveltime_list = []
        for travel_info in traveltime_list[detour_info]:
            if start_time is None: break

            if travel_info["remove_time"] > start_time:

                tmp_traveltime_list.append(travel_info["life_time"])

        sum_traveltime = sum(tmp_traveltime_list)
        
        if len(tmp_traveltime_list) == 0:
            cnt = 1
        else:
            cnt = len(tmp_traveltime_list)
        
        return sum_traveltime / cnt

    #ある時間以降の消滅時間を持つ車両をカウント
    def count_traveltime(traveltime_list,detour_info,start_time):
        cnt = 0
        for travel_info in traveltime_list[detour_info]:
            if start_time is None: break

            if travel_info["remove_time"] > start_time:

                cnt += 1

        return cnt

    highway_ave = average_traveltime(traveltime_list,"highway")
    general_ave = average_traveltime(traveltime_list,"general")
    detour_ave  = average_traveltime(traveltime_list,"detour")

    highway_400ave = average_traveltime(traveltime_list,"highway",highway.start_control_time)
    general_400ave = average_traveltime(traveltime_list,"general",highway.start_control_time)
    detour_400ave  = average_traveltime(traveltime_list,"detour" ,highway.start_control_time)

    highway_400cnt = count_traveltime(traveltime_list,"highway",highway.start_control_time)
    general_400cnt = count_traveltime(traveltime_list,"general",highway.start_control_time)
    detour_400cnt  = count_traveltime(traveltime_list,"detour" ,highway.start_control_time)

    #平均旅行時間表示
    ws.cell(row = 1, column = 6).value=("平均旅行時間")
    ws.cell(row = 2, column = 6).value=("一般道路発生時間交通量")
    ws.cell(row = 3, column = 6).value=("高速道路のみ走行車両平均旅行時間")
    ws.cell(row = 4, column = 6).value=("一般道路のみ走行車両平均旅行時間")
    ws.cell(row = 5, column = 6).value=("迂回車両平均旅行時間")

    ws.cell(row = 2, column = 7).value=(general.queue[1]*3600/generate_time_max*10)
    ws.cell(row = 3, column = 7).value=(highway_ave)
    ws.cell(row = 4, column = 7).value=(general_ave)
    ws.cell(row = 5, column = 7).value=(detour_ave)

    ws.cell(row = 7, column = 6).value=("高速道路のみ走行車両台数")
    ws.cell(row = 8, column = 6).value=("一般道路のみ走行車両台数")
    ws.cell(row = 9, column = 6).value=("迂回車両台数")

    ws.cell(row = 7, column = 7).value=(len(traveltime_list["highway"]))
    ws.cell(row = 8, column = 7).value=(len(traveltime_list["general"]))
    ws.cell(row = 9, column = 7).value=(len(traveltime_list["detour"]))

    #400秒以降平均旅行時間表示
    ws.cell(row = 1, column = 9).value=("迂回制御開始以降平均旅行時間")
    ws.cell(row = 2, column = 9).value=("一般道路発生時間交通量")
    ws.cell(row = 3, column = 9).value=("高速道路のみ走行車両平均旅行時間")
    ws.cell(row = 4, column = 9).value=("一般道路のみ走行車両平均旅行時間")
    ws.cell(row = 5, column = 9).value=("迂回車両平均旅行時間")

    ws.cell(row = 2, column = 10).value=(general.queue[1]*3600/generate_time_max*10)
    ws.cell(row = 3, column = 10).value=(highway_400ave)
    ws.cell(row = 4, column = 10).value=(general_400ave)
    ws.cell(row = 5, column = 10).value=(detour_400ave)

    ws.cell(row = 7, column = 9).value=("高速道路のみ走行車両台数")
    ws.cell(row = 8, column = 9).value=("一般道路のみ走行車両台数")
    ws.cell(row = 9, column = 9).value=("迂回車両台数")

    ws.cell(row = 7, column = 10).value=(highway_400cnt)
    ws.cell(row = 8, column = 10).value=(general_400cnt)
    ws.cell(row = 9, column = 10).value=(detour_400cnt)

def traveltime_info_sheet(wb,traveltime_list):
    title = "旅行時間情報"
    ws = wb.create_sheet(title=title)

    #旅行時間出力
    def output_traveltime(traveltime_list,detour_info,start_row,start_col):
        ws.cell(row = start_row, column = start_col  ).value=("車両ID")
        ws.cell(row = start_row, column = start_col+1).value=("生成時間")
        ws.cell(row = start_row, column = start_col+2).value=("消滅時間")
        ws.cell(row = start_row, column = start_col+3).value=("生存時間")
        
        row_num = start_row + 1
        for travel_info in traveltime_list[detour_info]:

            ws.cell(row = row_num, column = start_col  ).value=(travel_info["car_id"])
            ws.cell(row = row_num, column = start_col+1).value=(travel_info["generate_time"])
            ws.cell(row = row_num, column = start_col+2).value=(travel_info["remove_time"])
            ws.cell(row = row_num, column = start_col+3).value=(travel_info["life_time"])
            row_num += 1
    
    ws.cell(row = 1, column = 1 ).value=("highway")
    output_traveltime(  traveltime_list,
                        detour_info="highway",
                        start_row = 2,
                        start_col = 1)

    ws.cell(row = 1, column = 6 ).value=("general")
    output_traveltime(  traveltime_list,
                        detour_info="general",
                        start_row = 2,
                        start_col = 6)

    ws.cell(row = 1, column = 11).value=("detour")
    output_traveltime(  traveltime_list,
                        detour_info="detour",
                        start_row = 2,
                        start_col = 11)

def create_visual_sheet(wb,car_log,car_startID,road_id,time_max,interval):
    #書式設定
    border = Border(left=Side(style='thin', color='000000'))
    fill = PatternFill(patternType='solid', fgColor='000000')

    title = "可視化 道路ID" + str(road_id)
    ws = wb.create_sheet(title=title)

    #車線数
    lane_cnt = 3

    for i in range(time_max):
        #各列の幅を指定
        col=abc_from_number(i+1)
        ws.column_dimensions[col].width=8

        #sheetの1行目に時間を表示
        if i%interval==0:
            ws.cell(row=1, column=1+lane_cnt*i/interval).value =(i/interval)
            ws.cell(row=1, column=1+lane_cnt*i/interval).border=border

    #if条件を満たすcar_logを出力
    for car in car_log:
        if (car.time%interval == 0 and
            car.front   >  0 and 
            car.lane    >  0 and
            car.road_id == road_id
            ):

            #car_idの下一桁を参照して色を付ける
            color_tmp=colorBarRGB(car.car_id)
            ws.cell(row=car.front, column=1+car.lane+lane_cnt*car.time/interval).font=Font(b=True,color=color_tmp)
            #IDと速度を記録
            ws.cell(row=car.front, column=1+car.lane+lane_cnt*car.time/interval).value=(str(car.car_id)+str("/")+str(int(car.vel*3.6)))

            car_id = car.car_id - car_startID
            #車線が変わった時だけセル色を変える
            if car.shift_lane != 0:
                ws.cell(row=car.front, column=1+car.lane+lane_cnt*car.time/interval).fill=fill
    
    #先頭行固定
    ws.freeze_panes = 'A2'

def create_log_sheet(wb,car_log,road_id,interval):
    title = "log 道路ID" + str(road_id)
    ws = wb.create_sheet(title=title)

    #sheetの1行目に列名を書き込む
    col_title= ["時間",
                "車両ID",
                "前方位置",
                "後方位置",
                "道路ID",
                "車線",
                "速度",
                "加速度",
                "車間",
                "相対速度",
                "前方車両ID",
                "車線変更途中",
                "車線変更先",
                "車線変更開始時間",
                "車線変更先車間距離",
                "車線変更先前方車両ID",
                "目標車両ID",
                "車両type",
                "渋滞予見フラグ",
                "JADフラグ",
                "迂回情報"
                ]
    for col_loc,title in enumerate(col_title,1):#1行目の文字の部分の書き込み col_tmpは1からスタートする
        ws.cell(row=1, column=col_loc).value=(title)

    #先頭行固定
    ws.freeze_panes = 'A2'

    #2行目以降にif条件を満たすcar_logを出力
    output_cnt = 0
    for car in car_log:
        if (car.time%interval == 0 and
            # car.lane    != 0 and
            car.road_id == road_id
            ):
            
            #出力するパラメータをリストへ(SimpleNamespaceはPython3.9以降じゃないと順番が保証されないため作成)
            log_list = [car.time/10,
                        car.car_id,
                        car.front,
                        car.back,
                        car.road_id,
                        car.lane,
                        car.vel,
                        car.accl,
                        car.inter_vehicle_distance,
                        car.delta_vel,
                        car.front_car_id,
                        car.shift_lane,
                        car.shift_lane_to,
                        car.shift_begin,
                        car.shift_distance_go,
                        car.shift_front_car_id,
                        car.target_id,
                        car.type,
                        car.jam_prevision_flag,
                        car.jam_absorption_flag,
                        car.detour_info]
            
            for col_loc,value in enumerate(log_list,1):
                ws.cell(row=2+output_cnt, column=col_loc).value=(value)
            
            output_cnt+=1

#ファイルの名前を決め、ファイルを作成する
def create_path(seed,ratio_type,jam_prevision_vel,JAD_active,target_vol,avoiding_time):
    now = datetime.datetime.now()#現在時刻取得
    #日付を0埋め(ファイルの順番的に埋めたほうが楽)
    year_0p   = str(now.year  ).zfill(4)
    month_0p  = str(now.month ).zfill(2)
    day_0p    = str(now.day   ).zfill(2)
    hour_0p   = str(now.hour  ).zfill(2)
    minute_0p = str(now.minute).zfill(2)
    second_0p = str(now.second).zfill(2)
    microsecond_0p = str(now.microsecond).zfill(6)
    #201910301234みたいに作る
    date = year_0p + month_0p + day_0p + hour_0p + minute_0p + second_0p + microsecond_0p
    
    save_dir = os.getcwd()+"/../output"
    os.makedirs(save_dir, exist_ok=True)
    path = save_dir + "/seed" + str(seed).zfill(3) \
                    + "_" + "車載率" + str(ratio_type[0]) \
                    + "_" + "速度閾値" + str(jam_prevision_vel).zfill(3) \
                    + "_" + "目標交通量" + str(target_vol) \
                    + "_" + "迂回制御時間" + str(avoiding_time).zfill(3) + "s" \
                    + "_" + str(date) + ".xlsx"
    return path

#excel直接出力（加速・走行・追い越し用）
def save( log,
          traveltime_list,
          car_max,
          time_max,
          generate_time_max,
          interval,
          road_list,
          seed,
          prc_time,
          ratio_type,
          jam_prevision_vel,
          JAD_active,
          target_vol,
          avoiding_time
          ):

    path=create_path(seed,ratio_type,jam_prevision_vel,JAD_active,target_vol,avoiding_time)

    wb,ws=create_excel_file()
    ws.title ="情報"#最初は'sheet1'というシートがデフォルトで作成されるため、名前変更
    create_info_sheet(ws,time_max,car_max,road_list,prc_time)
    detour_info_sheet(wb,log,traveltime_list,road_list,time_max,generate_time_max,interval)
    traveltime_info_sheet(wb,traveltime_list)
    
    for road_id in range(len(road_list)):
        create_visual_sheet(wb,log,road_list[road_id].start_carID,road_id,time_max,interval)
        create_log_sheet(wb,log,road_id,interval)

    wb.save(path)
    print("Complete",path)
